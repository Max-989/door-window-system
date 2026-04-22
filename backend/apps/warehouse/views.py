"""
warehouse app - 视图
"""
import openpyxl
from django_filters import rest_framework as filters
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from common.responses import error

from .models import (
    AccessoryInventory,
    HardwareInventory,
    PendingGoods,
    WarehouseProduct,
    WarehouseTransfer,
)
from .serializers import (
    AccessoryInventorySerializer,
    BulkImportSerializer,
    HardwareInventorySerializer,
    PendingGoodsSerializer,
    StockRecordSerializer,
    StocktakeSerializer,
    WarehouseProductSerializer,
    WarehouseTransferSerializer,
)
from .services import StockService


class WarehouseProductFilterSet(filters.FilterSet):
    """产品流转仓筛选"""

    city = filters.CharFilter(field_name="city", lookup_expr="icontains")
    date_start = filters.DateTimeFilter(field_name="created_at", lookup_expr="gte")
    date_end = filters.DateTimeFilter(field_name="created_at", lookup_expr="lte")
    # 批次1新增筛选参数
    order_status = filters.CharFilter(field_name="order_status", help_text="订单状态")
    brand_id = filters.NumberFilter(field_name="order__brand", help_text="品牌ID")
    store_id = filters.NumberFilter(field_name="order__store", help_text="门店ID")
    auto_generated = filters.BooleanFilter(
        field_name="auto_generated", help_text="是否自动生成"
    )

    class Meta:
        model = WarehouseProduct
        fields = [
            "order_no",
            "product_type",
            "status",
            "city",
            "order_status",
            "brand_id",
            "store_id",
            "auto_generated",
        ]


class WarehouseProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = WarehouseProduct.objects.select_related(
        "operator", "order", "order__brand", "order__store"
    ).all()
    serializer_class = WarehouseProductSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = WarehouseProductFilterSet
    search_fields = ["order_no", "product_model"]


class HardwareInventoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = HardwareInventory.objects.all()
    serializer_class = HardwareInventorySerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["hardware_type", "city"]
    search_fields = ["name"]

    @action(detail=True, methods=["post"], url_path="out-stock")
    def out_stock(self, request, pk=None):
        """出库"""
        error_response, item = StockService.out_stock(
            model_class=HardwareInventory,
            item_type="hardware",
            pk=pk,
            quantity=request.data.get("quantity"),
            reason=request.data.get("reason", ""),
            related_task_id=request.data.get("related_task_id"),
            operator=request.user,
        )
        if error_response:
            return error_response

        result = HardwareInventorySerializer(item).data
        if item.current_stock < item.alert_quantity:
            result["warning"] = f"库存低于预警数量（{item.alert_quantity}）"

        return Response(result)

    @action(detail=True, methods=["post"], url_path="in-stock")
    def in_stock(self, request, pk=None):
        """入库"""
        error_response, item = StockService.in_stock(
            model_class=HardwareInventory,
            item_type="hardware",
            pk=pk,
            quantity=request.data.get("quantity"),
            reason=request.data.get("reason", ""),
            supplier=request.data.get("supplier", ""),
            operator=request.user,
        )
        if error_response:
            return error_response

        result = HardwareInventorySerializer(item).data
        if item.current_stock < item.alert_quantity:
            result["warning"] = f"库存仍低于预警数量（{item.alert_quantity}）"

        return Response(result)

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """批量导入五金库存"""
        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data["file"]
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created_count = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or not row[0]:
                    continue
                HardwareInventory.objects.create(
                    name=str(row[0]).strip(),
                    hardware_type=str(row[1]).strip() if len(row) > 1 else "",
                    current_stock=int(row[2]) if len(row) > 2 and row[2] else 0,
                    alert_quantity=int(row[3]) if len(row) > 3 and row[3] else 0,
                    city=str(row[4]).strip() if len(row) > 4 else "",
                )
                created_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        return Response(
            {"created": created_count, "errors": errors}, status=status.HTTP_200_OK
        )


class AccessoryInventoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = AccessoryInventory.objects.all()
    serializer_class = AccessoryInventorySerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["city"]
    search_fields = ["name", "spec_model"]

    @action(detail=True, methods=["post"], url_path="out-stock")
    def out_stock(self, request, pk=None):
        """出库"""
        error_response, item = StockService.out_stock(
            model_class=AccessoryInventory,
            item_type="accessory",
            pk=pk,
            quantity=request.data.get("quantity"),
            reason=request.data.get("reason", ""),
            related_task_id=request.data.get("related_task_id"),
            operator=request.user,
        )
        if error_response:
            return error_response

        return Response(AccessoryInventorySerializer(item).data)

    @action(detail=True, methods=["post"], url_path="in-stock")
    def in_stock(self, request, pk=None):
        """入库"""
        error_response, item = StockService.in_stock(
            model_class=AccessoryInventory,
            item_type="accessory",
            pk=pk,
            quantity=request.data.get("quantity"),
            reason=request.data.get("reason", ""),
            supplier=request.data.get("supplier", ""),
            operator=request.user,
        )
        if error_response:
            return error_response

        return Response(AccessoryInventorySerializer(item).data)

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """批量导入配件库存"""
        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data["file"]
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created_count = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or not row[0]:
                    continue
                AccessoryInventory.objects.create(
                    name=str(row[0]).strip(),
                    spec_model=str(row[1]).strip() if len(row) > 1 else "",
                    current_stock=int(row[2]) if len(row) > 2 and row[2] else 0,
                    city=str(row[3]).strip() if len(row) > 3 else "",
                )
                created_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        return Response(
            {"created": created_count, "errors": errors}, status=status.HTTP_200_OK
        )


class PendingGoodsViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = PendingGoods.objects.all()
    serializer_class = PendingGoodsSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["door_type", "match_status", "city"]
    search_fields = ["goods_name"]

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """批量导入货品暂存"""
        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data["file"]
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created_count = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or not row[0]:
                    continue
                PendingGoods.objects.create(
                    goods_name=str(row[0]).strip(),
                    door_type=str(row[1]).strip() if len(row) > 1 else "",
                    color=str(row[2]).strip() if len(row) > 2 else "",
                    city=str(row[3]).strip() if len(row) > 3 else "",
                )
                created_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        return Response(
            {"created": created_count, "errors": errors}, status=status.HTTP_200_OK
        )


class WarehouseTransferViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    """城市间调拨 CRUD"""
    queryset = WarehouseTransfer.objects.select_related(
        "initiated_by", "confirmed_by"
    ).all()
    serializer_class = WarehouseTransferSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["from_city", "to_city", "status"]
    search_fields = ["from_city", "to_city"]
    ordering_fields = ["created_at"]

    @action(detail=True, methods=["patch"], url_path="update-status")
    def update_status(self, request, pk=None):
        """更新调拨状态"""
        transfer = self.get_object()
        new_status = request.data.get("status")
        if not new_status:
            return error(message="status字段必填", code=status.HTTP_400_BAD_REQUEST)
        valid_statuses = ["pending", "confirmed", "transit", "completed"]
        if new_status not in valid_statuses:
            return error(message=f"无效状态，可选: {valid_statuses}", code=status.HTTP_400_BAD_REQUEST)
        transfer.status = new_status
        if new_status in ("confirmed", "completed") and request.user.is_authenticated:
            transfer.confirmed_by = request.user
        transfer.save(update_fields=["status", "confirmed_by", "updated_at"])
        return Response(WarehouseTransferSerializer(transfer).data)


class StocktakeView(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    """盘点接口"""

    def create(self, request):
        """盘点：更新库存并记录差异"""
        serializer = StocktakeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        warehouse_type = data["warehouse_type"]
        item_id = data["item_id"]
        actual_qty = data["actual_quantity"]

        model_map = {
            "hardware": (HardwareInventory, "current_stock"),
            "accessory": (AccessoryInventory, "current_stock"),
            "pending": (PendingGoods, "current_stock")
            if hasattr(PendingGoods, "current_stock")
            else None,
        }

        if warehouse_type not in model_map:
            return error(message="无效仓库类型，可选: hardware/accessory", code=status.HTTP_400_BAD_REQUEST)

        model_class, stock_field = model_map[warehouse_type]
        if model_class is None:
            return error(message="该仓库类型暂不支持盘点", code=status.HTTP_400_BAD_REQUEST)

        try:
            item = model_class.objects.get(id=item_id)
        except model_class.DoesNotExist:
            return error(message="物品不存在", code=status.HTTP_404_NOT_FOUND)

        system_qty = getattr(item, stock_field, 0)
        difference = actual_qty - system_qty

        # 更新库存
        setattr(item, stock_field, actual_qty)
        # 如果有 available_stock 字段也同步更新
        if hasattr(item, "available_stock") and stock_field == "current_stock":
            item.available_stock = actual_qty - getattr(item, "pending_out_quantity", 0)
        item.save()

        return Response(
            {
                "item_id": item_id,
                "warehouse_type": warehouse_type,
                "item_name": str(item),
                "system_quantity": system_qty,
                "actual_quantity": actual_qty,
                "difference": difference,
            }
        )
